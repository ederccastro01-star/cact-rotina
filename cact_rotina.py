# -*- coding: utf-8 -*-
"""
CACT - Rotina de Publicações Processuais
========================================
Castro Advocacia & Cálculos Trabalhistas

Rotina que replica o aplicativo CACT e a concilia com o cruzamento da
"CACT - Agenda de Prazos". Executa de segunda a sexta às 5h30 (via Agendador
de Tarefas do Windows) e:

  1. Busca as publicações do dia no DJEN (API Comunica/PJe do CNJ) pelos
     advogados (nome + OAB) do escritório e, opcionalmente, por uma lista de
     processos.
  2. Deduplica e gera o "relatorio-detalhado-castro (DD.MM.AAAA).xlsx".
  3. Cruza os números de processo com as guias "Publicações" e "Prazos Feitos"
     da Agenda de Prazos.
  4. Para os processos que cruzam, classifica a peça, calcula o prazo,
     destaca "Atenção! Honorários", agenda "Embargos de Declaração" (5 dias) e
     escreve o resumo (coluna L) — usando a API do Claude (Anthropic) quando há
     chave configurada, ou uma classificação heurística como alternativa.
  5. Gera o "CACT - Relatório Cruzamento Publicações DD-MM-AAAA.xlsx" no mesmo
     formato da guia "Publicações".

Uso:
    python cact_rotina.py                      # fonte DJEN, data = hoje
    python cact_rotina.py --data 23.06.2026    # força a data de disponibilização
    python cact_rotina.py --fonte arquivo --arquivo "caminho/relatorio.xlsx"
    python cact_rotina.py --sem-ia             # ignora a API do Claude (só heurística)

Configuração: edite o arquivo config.json (mesma pasta deste script).
"""

import os
import re
import sys
import json
import time
import html
import argparse
import unicodedata
from datetime import date, datetime, timedelta

try:
    import requests
except ImportError:
    requests = None

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --------------------------------------------------------------------------- #
# Utilidades
# --------------------------------------------------------------------------- #
AQUI = os.path.dirname(os.path.abspath(__file__))


def log(msg):
    print(f"[{datetime.now():%d/%m/%Y %H:%M:%S}] {msg}", flush=True)


def carregar_config():
    caminho = os.path.join(AQUI, "config.json")
    if not os.path.exists(caminho):
        log("ERRO: config.json não encontrado ao lado do script.")
        sys.exit(1)
    with open(caminho, "r", encoding="utf-8") as fh:
        cfg = json.load(fh)
    # Permite sobrepor a chave por variável de ambiente (mais seguro)
    cfg["anthropic_api_key"] = (
        os.environ.get("ANTHROPIC_API_KEY")
        or cfg.get("anthropic_api_key", "")
    ).strip()
    return cfg


def so_digitos(s):
    return re.sub(r"\D", "", str(s)) if s is not None else ""


def norm_proc(s):
    """Normaliza número de processo para comparação (somente dígitos, >=15)."""
    d = so_digitos(s)
    return d if len(d) >= 15 else None


def sem_acento(s):
    s = str(s or "")
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


# --------------------------------------------------------------------------- #
# Datas / dias úteis
# --------------------------------------------------------------------------- #
def carregar_feriados(cfg):
    fer = set()
    for d in cfg.get("feriados", []):
        try:
            fer.add(datetime.strptime(d, "%Y-%m-%d").date())
        except ValueError:
            pass
    return fer


def eh_util(d, feriados):
    return d.weekday() < 5 and d not in feriados


def proximo_dia_util(d, feriados):
    d = d + timedelta(days=1)
    while not eh_util(d, feriados):
        d += timedelta(days=1)
    return d


def prazo_em_dias_uteis(base, n, feriados):
    """Retorna a data correspondente ao N-ésimo dia útil após 'base'
    (a contagem começa no primeiro dia útil seguinte a 'base')."""
    d = base
    c = 0
    while c < n:
        d += timedelta(days=1)
        if eh_util(d, feriados):
            c += 1
    return d


WD_PT = {0: "seg", 1: "ter", 2: "qua", 3: "qui", 4: "sex", 5: "sáb", 6: "dom"}


# --------------------------------------------------------------------------- #
# 1) Captura das publicações
# --------------------------------------------------------------------------- #
# Padrões para classificar automaticamente cada item do cadastro
RE_PROCESSO_MASC = re.compile(r"\d{7}-?\d{2}\.?\d{4}\.?\d\.?\d{2}\.?\d{4}")
RE_OAB = re.compile(r"^\s*(?:OAB[\s:/.\-]*)?([A-Za-z]{2})[\s:/.\-]*0*(\d{2,6})\s*$")
# Cabeçalho = a célula é APENAS um rótulo de coluna (sem dados em seguida)
RE_CABECALHO = re.compile(r"^\s*(advogad\w*|oab|processos?|n[uú]meros?|nomes?|cadastro)\s*$", re.I)


def _parse_oab(texto):
    """Devolve (uf, numero) se a célula parecer uma OAB; senão None."""
    m = RE_OAB.match(str(texto))
    if not m:
        return None
    uf = m.group(1).upper()
    num = m.group(2)
    # evita confundir com palavra comum de 2 letras + número solto
    return (uf, num)


def _classificar_item(valor, uf_padrao):
    """Classifica uma célula como ('processo'|'oab'|'nome', dado)."""
    s = str(valor).strip()
    if not s or RE_CABECALHO.match(s):
        return None
    digitos = so_digitos(s)
    # processo: máscara CNJ ou 20 dígitos
    if RE_PROCESSO_MASC.search(s) or len(digitos) >= 18:
        return ("processo", s.strip())
    # OAB com UF (ex.: GO14725, GO-0014725, OAB/GO 14725)
    oab = _parse_oab(s)
    if oab:
        return ("oab", {"uf": oab[0], "oab": str(int(oab[1]))})
    # número solto curto (provável OAB sem UF) -> usa UF padrão
    if s.replace(" ", "").isdigit() and 3 <= len(digitos) <= 6:
        return ("oab", {"uf": uf_padrao, "oab": str(int(digitos))})
    # tem letras -> nome de advogado
    if re.search(r"[A-Za-zÀ-ÿ]", s):
        return ("nome", re.sub(r"\s+", " ", s).strip())
    return None


def carregar_cadastro(cfg):
    """Lê o arquivo de cadastro (CACT - Cadastro Publicações) de forma flexível,
    varrendo TODAS as células de TODAS as abas (ou da aba configurada) e
    classificando cada item como nome de advogado, OAB ou número de processo.

    Assim, o escritório pode acrescentar novos nomes, OABs ou processos a qualquer
    momento — em qualquer linha/coluna/aba — sem precisar alterar a rotina."""
    caminho = cfg.get("arquivo_cadastro") or cfg.get("arquivo_advogados")
    aba_cfg = cfg.get("aba_cadastro")           # None = todas as abas
    uf_padrao = cfg.get("uf_padrao", "GO")
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    abas = [aba_cfg] if (aba_cfg and aba_cfg in wb.sheetnames) else wb.sheetnames

    nomes, oabs, processos = set(), {}, set()
    for nome_aba in abas:
        for row in wb[nome_aba].iter_rows(values_only=True):
            for cel in (row or ()):
                if cel is None:
                    continue
                cls = _classificar_item(cel, uf_padrao)
                if not cls:
                    continue
                tipo, dado = cls
                if tipo == "processo":
                    processos.add(dado)
                elif tipo == "oab":
                    oabs[(dado["uf"], dado["oab"])] = dado
                elif tipo == "nome":
                    nomes.add(dado)
    wb.close()

    cadastro = {
        "nomes": sorted(nomes),
        "oabs": list(oabs.values()),
        "processos": sorted(processos),
    }
    log(f"Cadastro carregado de '{os.path.basename(caminho)}': "
        f"{len(cadastro['oabs'])} OABs, {len(cadastro['nomes'])} nomes, "
        f"{len(cadastro['processos'])} processos.")
    return cadastro


def _extrai_campo(item, *chaves):
    for k in chaves:
        if k in item and item[k] not in (None, ""):
            return item[k]
    return ""


def _texto_item(item):
    txt = _extrai_campo(item, "texto", "teor", "conteudo", "textoComunicacao")
    txt = html.unescape(str(txt))
    txt = re.sub(r"<[^>]+>", " ", txt)            # remove HTML
    txt = re.sub(r"[ \t]+", " ", txt).strip()
    return txt


def _processo_item(item):
    return _extrai_campo(
        item, "numeroprocessocommascara", "numero_processo",
        "numeroProcesso", "numeroProcessoComMascara", "processo"
    )


def buscar_djen(cfg, data_disp, cadastro, processos_extra):
    """Consulta a API Comunica/DJEN. Retorna lista de dicts {processo, texto, item}."""
    if requests is None:
        log("ERRO: biblioteca 'requests' não instalada. Rode Instalar.bat.")
        sys.exit(1)

    base = cfg.get("djen_url", "https://comunicaapi.pje.jus.br/api/v1/comunicacao")
    data_str = data_disp.strftime("%Y-%m-%d")
    headers = {"User-Agent": "CACT-Rotina/1.0", "Accept": "application/json"}
    timeout = cfg.get("timeout_segundos", 60)
    pausa = cfg.get("pausa_entre_consultas_seg", 0.5)
    itens_por_pagina = cfg.get("itens_por_pagina", 100)

    coletados = {}     # chave de dedupe -> dict
    def adiciona(item):
        proc = _processo_item(item)
        texto = _texto_item(item)
        if not texto:
            return
        ident = _extrai_campo(item, "id", "hash") or (so_digitos(proc) + "|" + texto[:120])
        if ident in coletados:
            return
        coletados[ident] = {"processo": proc, "texto": texto, "item": item}

    def consulta(params, descricao):
        pagina = 1
        while True:
            p = dict(params)
            p["pagina"] = pagina
            p["itensPorPagina"] = itens_por_pagina
            try:
                r = requests.get(base, params=p, headers=headers, timeout=timeout)
                if r.status_code != 200:
                    log(f"  ! {descricao}: HTTP {r.status_code}")
                    return
                dados = r.json()
            except Exception as e:
                log(f"  ! {descricao}: erro {e}")
                return
            itens = dados.get("items") or dados.get("result") or dados.get("comunicacoes") or []
            for it in itens:
                adiciona(it)
            total = dados.get("count") or dados.get("total") or len(itens)
            if len(itens) < itens_por_pagina or pagina * itens_por_pagina >= int(total or 0):
                break
            pagina += 1
            time.sleep(pausa)

    base_data = {"dataDisponibilizacaoInicio": data_str, "dataDisponibilizacaoFim": data_str}

    # 1) Busca por OAB (mais confiável e abrangente)
    log(f"Consultando DJEN — disponibilização {data_disp:%d/%m/%Y} ...")
    for o in cadastro.get("oabs", []):
        consulta(dict(base_data, numeroOab=o["oab"], ufOab=o["uf"]),
                 f"OAB {o['uf']} {o['oab']}")
        time.sleep(pausa)

    # 2) Busca por nome de advogado (opcional — captura quem não tem OAB no cadastro)
    if cfg.get("buscar_por_nome", True):
        for nome in cadastro.get("nomes", []):
            consulta(dict(base_data, nomeAdvogado=nome), f"Nome {nome}")
            time.sleep(pausa)

    # 3) Busca por número de processo (do cadastro + lista extra do config)
    todos_proc = list(dict.fromkeys(list(cadastro.get("processos", [])) + list(processos_extra)))
    for proc in todos_proc:
        consulta(dict(base_data, numeroProcesso=proc), f"Processo {proc}")
        time.sleep(pausa)

    log(f"Publicações capturadas (após dedupe): {len(coletados)}")
    return list(coletados.values())


def carregar_de_arquivo(caminho):
    """Modo alternativo: lê um 'relatorio-detalhado-castro' já exportado."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    pubs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        proc = str(row[0]).strip()
        texto = html.unescape(str(row[1] or ""))
        texto = re.sub(r"<[^>]+>", " ", texto)
        texto = re.sub(r"[ \t]+", " ", texto).strip()
        pubs.append({"processo": proc, "texto": texto, "item": {}})
    wb.close()
    log(f"Publicações lidas do arquivo: {len(pubs)}")
    return pubs


def salvar_relatorio_detalhado(pubs, cfg, data_disp):
    pasta = cfg.get("pasta_historico", cfg["pasta_saida"])
    os.makedirs(pasta, exist_ok=True)
    nome = f"relatorio-detalhado-castro ({data_disp:%d.%m.%Y}).xlsx"
    caminho = os.path.join(pasta, nome)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "relatorio-detalhado-castro"
    ws.append(["Número do processo", "Inteiro teor da intimação"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for p in pubs:
        ws.append([p["processo"], p["texto"]])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120
    wb.save(caminho)
    log(f"Relatório detalhado salvo: {nome}")
    return caminho


# --------------------------------------------------------------------------- #
# 2) Cruzamento com a Agenda de Prazos
# --------------------------------------------------------------------------- #
def indexar_agenda(cfg):
    """Lê as guias Publicações e Prazos Feitos; devolve:
    - mapa norm_proc -> dados de referência (última ocorrência preferindo Publicações)
    - conjuntos de processos de cada guia."""
    caminho = cfg["arquivo_agenda"]
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ref_pub, ref_pf = {}, {}
    set_pub, set_pf = set(), set()

    def varre(aba, store, conj):
        if aba not in wb.sheetnames:
            return
        for row in wb[aba].iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 11:
                continue
            n = norm_proc(row[4]) if len(row) > 4 else None
            if not n:
                continue
            conj.add(n)
            store[n] = {
                "Reclamante": row[5], "Reclamado": row[6], "Advogado": row[7],
                "Prioridade": row[8], "Complexidade": row[9], "Despacho": row[10],
            }

    varre(cfg.get("aba_publicacoes", "Publicações"), ref_pub, set_pub)
    varre(cfg.get("aba_prazos_feitos", "Prazos Feitos"), ref_pf, set_pf)
    wb.close()

    ref = {}
    for n in set_pub | set_pf:
        ref[n] = {
            "ref": ref_pub.get(n) or ref_pf.get(n) or {},
            "inPub": n in set_pub, "inPF": n in set_pf,
        }
    log(f"Agenda indexada — Publicações: {len(set_pub)} | Prazos Feitos: {len(set_pf)}")
    return ref


def cruzar(pubs, ref):
    """Retorna apenas as publicações cujo processo consta na Agenda, sem repetir processo."""
    vistos = set()
    matches = []
    for p in pubs:
        n = norm_proc(p["processo"])
        if n and n in ref and n not in vistos:
            vistos.add(n)
            mp = dict(p)
            mp["norm"] = n
            mp["ref"] = ref[n]["ref"]
            mp["inPub"] = ref[n]["inPub"]
            mp["inPF"] = ref[n]["inPF"]
            matches.append(mp)
    log(f"Processos que cruzaram: {len(matches)}")
    return matches


# --------------------------------------------------------------------------- #
# 3) Classificação (peça / prazo / honorários / ED / resumo)
# --------------------------------------------------------------------------- #
RE_PRAZO = re.compile(r"prazo\s*(?:comum\s*)?(?:de\s*)?(\d{1,3})\s*\(?\s*\w*\s*\)?\s*dias", re.I)
RE_PRAZO2 = re.compile(r"(\d{1,3})\s*dias", re.I)
RE_48H = re.compile(r"48\s*horas", re.I)

KW_HONORARIOS = [
    "pagamento", "liberação", "liberacao", "levantamento", "alvará", "alvara",
    "transferência", "transferencia", "transferida", "devolução de saldo",
    "certidão de crédito", "certidao de credito", "habilitação de crédito",
    "habilitacao de credito", "honorários", "honorarios", "crédito líquido",
    "credito liquido", "distribuição do numerário", "distribuicao do numerario",
]
KW_DECISAO = ["sentença", "sentenca", "acórdão", "acordao", "decisão", "decisao",
              "homologo os cálculos", "homologo os calculos", "dispositivo"]


def classificar_heuristica(texto, feriados=None, data_publicado=None):
    """Classificação por regras — usada apenas quando NÃO há chave do Claude.
    É propositalmente conservadora: só agenda prazo para Embargos de Declaração
    (5 dias, ou o prazo menor explícito), evitando datas indevidas. Os prazos de
    manifestação/contrarrazões etc. devem ser conferidos manualmente ou via IA."""
    t = sem_acento(texto).lower()
    honor = any(sem_acento(k).lower() in t for k in KW_HONORARIOS)
    decisao = any(sem_acento(k).lower() in t for k in KW_DECISAO)

    # prazo explícito pequeno (apenas para limitar o ED a um prazo menor)
    prazo_expl = None
    if RE_48H.search(texto):
        prazo_expl = 2
    else:
        m = RE_PRAZO.search(texto) or RE_PRAZO2.search(texto)
        if m:
            try:
                v = int(m.group(1))
                if 1 <= v <= 15:        # ignora números grandes/ruído
                    prazo_expl = v
            except ValueError:
                pass

    prazo_dias = None
    if honor and decisao:
        peca = "Atenção! Honorários / Embargos de Declaração"
        prazo_dias = min(5, prazo_expl) if prazo_expl else 5
    elif honor:
        peca = "Atenção! Honorários"
    elif decisao:
        peca = "Embargos de Declaração"
        prazo_dias = min(5, prazo_expl) if prazo_expl else 5
    else:
        peca = "Acompanhamento"

    resumo = re.sub(r"\s+", " ", texto).strip()
    if len(resumo) > 600:
        resumo = resumo[:600].rsplit(" ", 1)[0] + " (...)"
    resumo += "  [Classificação automática por regras — confira o prazo no inteiro teor.]"
    if honor:
        resumo += " ATENÇÃO HONORÁRIOS."

    return {"peca": peca, "prazo_dias": prazo_dias, "resumo": resumo,
            "honorarios": honor, "decisao": decisao}


PROMPT_SISTEMA = """Você é assistente jurídico de um escritório de advocacia trabalhista (CACT - Castro Advocacia & Cálculos Trabalhistas). Recebe o inteiro teor de uma intimação/publicação processual e deve devolver uma classificação em JSON, seguindo ESTAS regras do escritório:

REGRAS:
1. "peca": a providência a ser tomada pelo escritório. Use rótulos curtos, por exemplo: "Acompanhamento", "Embargos de Declaração", "Manifestação sobre impugnação aos cálculos", "Contrarrazões", "Manifestação", "Informar dados/endereços", etc.
2. HONORÁRIOS: sempre que a publicação disser respeito a pagamento, liberação, levantamento, alvará, transferência de valores, devolução de saldo, certidão/habilitação de crédito ou honorários, defina "honorarios": true e a "peca" deve ser "Atenção! Honorários" (acrescentando, se houver outra providência, em seguida).
3. EMBARGOS DE DECLARAÇÃO: se a publicação for sentença, decisão ou acórdão de qualquer natureza, "peca" = "Embargos de Declaração" e "prazo_dias" = 5 (cinco), agendados ANTES do prazo de eventual recurso — EXCETO quando a publicação já previr prazo menor (use o menor).
4. "prazo_dias": número de dias ÚTEIS para cumprir a determinação dirigida AO ESCRITÓRIO (a parte que representamos). Se o prazo for da parte adversa, ou se for mero acompanhamento sem providência nossa, use null.
5. "resumo": resumo claro e objetivo (2 a 4 frases), em português simples, do teor da publicação e do que precisa ser feito; ao final, se for caso de honorários, inclua "ATENÇÃO HONORÁRIOS".

Responda SOMENTE com um objeto JSON válido, sem texto antes ou depois, no formato:
{"peca": "...", "prazo_dias": null ou inteiro, "honorarios": true/false, "decisao": true/false, "resumo": "..."}"""


def classificar_claude(texto, cfg):
    """Classificação com a API do Claude (Anthropic). Retorna dict ou None em caso de falha."""
    chave = cfg.get("anthropic_api_key", "")
    if not chave or requests is None:
        return None
    url = "https://api.anthropic.com/v1/messages"
    headers = {
        "x-api-key": chave,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    corpo = {
        "model": cfg.get("modelo_claude", "claude-sonnet-4-6"),
        "max_tokens": 700,
        "system": PROMPT_SISTEMA,
        "messages": [{"role": "user", "content": texto[:9000]}],
    }
    try:
        r = requests.post(url, headers=headers, json=corpo, timeout=cfg.get("timeout_segundos", 60))
        if r.status_code != 200:
            log(f"  ! Claude HTTP {r.status_code}: {r.text[:160]}")
            return None
        conteudo = r.json()["content"][0]["text"]
        m = re.search(r"\{.*\}", conteudo, re.S)
        if not m:
            return None
        dados = json.loads(m.group(0))
        return {
            "peca": dados.get("peca") or "Acompanhamento",
            "prazo_dias": dados.get("prazo_dias"),
            "resumo": dados.get("resumo") or "",
            "honorarios": bool(dados.get("honorarios")),
            "decisao": bool(dados.get("decisao")),
        }
    except Exception as e:
        log(f"  ! Claude erro: {e}")
        return None


# --------------------------------------------------------------------------- #
# 4) Geração do relatório de cruzamento
# --------------------------------------------------------------------------- #
HEADERS = ["Prazo", "Dia/Sem", "Peça", "Publicado", "Processo", "Reclamante",
           "Reclamado", "Advogado", "Prioridade", "Complexidade", "Despacho",
           "Resumo da Publicação"]


def gerar_relatorio_cruzamento(matches, cfg, data_disp, usar_ia):
    feriados = carregar_feriados(cfg)
    publicado = prazo_em_dias_uteis(data_disp, 1, feriados)  # disponibilização + 1 dia útil

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Publicações"

    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = hfill; c.font = hfont; c.border = borda
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hon_fill = PatternFill("solid", fgColor="FFF2CC")
    ed_fill = PatternFill("solid", fgColor="DDEBF7")

    r = 2
    for m in matches:
        cls = None
        if usar_ia:
            cls = classificar_claude(m["texto"], cfg)
        if cls is None:
            cls = classificar_heuristica(m["texto"], feriados, publicado)

        peca = cls["peca"]
        prazo = prazo_em_dias_uteis(publicado, cls["prazo_dias"], feriados) if cls.get("prazo_dias") else None
        rf = m["ref"]
        desp_old = rf.get("Despacho")
        desp_old = "" if desp_old in (None, "None") else str(desp_old)
        if sem_acento(desp_old).strip().lower().startswith("acompanhamento"):
            despacho_final = desp_old                      # já começa com a palavra
        else:
            despacho_final = ("Acompanhamento\n" + desp_old).strip()

        ws.cell(r, 1, prazo)
        ws.cell(r, 2, prazo)
        ws.cell(r, 3, peca)
        ws.cell(r, 4, publicado)
        ws.cell(r, 5, m["processo"])
        ws.cell(r, 6, rf.get("Reclamante"))
        ws.cell(r, 7, rf.get("Reclamado"))
        ws.cell(r, 8, rf.get("Advogado"))
        ws.cell(r, 9, rf.get("Prioridade"))
        ws.cell(r, 10, rf.get("Complexidade"))
        ws.cell(r, 11, despacho_final)
        ws.cell(r, 12, cls["resumo"])

        ws.cell(r, 1).number_format = "dd/mm/yy"
        ws.cell(r, 2).number_format = "ddd"
        ws.cell(r, 4).number_format = "dd/mm/yy"
        for j in range(1, 13):
            cel = ws.cell(r, j)
            cel.border = borda
            cel.alignment = Alignment(vertical="top", wrap_text=(j in (3, 6, 7, 8, 11, 12)))
        if "Honorários" in peca or cls.get("honorarios"):
            ws.cell(r, 3).fill = hon_fill
            ws.cell(r, 3).font = Font(bold=True, color="9C5700")
        elif "Embargos" in peca:
            ws.cell(r, 3).fill = ed_fill
        r += 1

    larguras = {"A": 10, "B": 7, "C": 24, "D": 10, "E": 24, "F": 26, "G": 28,
                "H": 24, "I": 12, "J": 13, "K": 40, "L": 60}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    os.makedirs(cfg["pasta_saida"], exist_ok=True)
    nome = f"CACT - Relatório Cruzamento Publicações {data_disp:%d-%m-%Y}.xlsx"
    caminho = os.path.join(cfg["pasta_saida"], nome)
    wb.save(caminho)
    log(f"Relatório de cruzamento salvo: {nome}  ({r-2} processos)")
    return caminho


# --------------------------------------------------------------------------- #
# Programa principal
# --------------------------------------------------------------------------- #
def executar_rotina(cfg, data_disp=None, fonte="djen", arquivo=None, sem_ia=False):
    """Executa toda a rotina a partir de um cfg já montado.
    Reutilizável tanto pelo modo local (main) quanto pela nuvem (runner_nuvem).
    Retorna a lista de caminhos de arquivos gerados."""
    feriados = carregar_feriados(cfg)
    if data_disp is None:
        data_disp = date.today()

    gerados = []

    # Em fim de semana/feriado a rotina não busca publicações.
    if fonte == "djen" and not eh_util(data_disp, feriados):
        log(f"{data_disp:%d/%m/%Y} não é dia útil — nada a fazer.")
        return gerados

    usar_ia = (not sem_ia) and bool(cfg.get("anthropic_api_key"))
    if not usar_ia and not sem_ia:
        log("AVISO: sem chave da API do Claude — usando classificação heurística. "
            "Configure 'anthropic_api_key' para resumos por IA.")

    # 1) captura
    if fonte == "arquivo":
        if not arquivo:
            log("ERRO: informe o arquivo no modo arquivo.")
            sys.exit(1)
        pubs = carregar_de_arquivo(arquivo)
    else:
        cadastro = carregar_cadastro(cfg)
        processos_extra = cfg.get("processos_monitorados", [])
        pubs = buscar_djen(cfg, data_disp, cadastro, processos_extra)
        if pubs:
            gerados.append(salvar_relatorio_detalhado(pubs, cfg, data_disp))

    if not pubs:
        log("Nenhuma publicação capturada. Encerrando.")
        return gerados

    # 2) cruzamento
    ref = indexar_agenda(cfg)
    matches = cruzar(pubs, ref)
    if not matches:
        log("Nenhum processo de interesse cruzou hoje. Encerrando.")
        return gerados

    # 3-4) classificação + relatório
    gerados.append(gerar_relatorio_cruzamento(matches, cfg, data_disp, usar_ia))
    log("Rotina concluída com sucesso.")
    return gerados


def main():
    ap = argparse.ArgumentParser(description="CACT - Rotina de Publicações")
    ap.add_argument("--data", help="Data de disponibilização (DD.MM.AAAA). Padrão: hoje")
    ap.add_argument("--fonte", choices=["djen", "arquivo"], default="djen")
    ap.add_argument("--arquivo", help="Caminho do relatorio-detalhado (modo arquivo)")
    ap.add_argument("--sem-ia", action="store_true", help="Não usar a API do Claude")
    args = ap.parse_args()

    cfg = carregar_config()
    data_disp = datetime.strptime(args.data, "%d.%m.%Y").date() if args.data else date.today()
    executar_rotina(cfg, data_disp=data_disp, fonte=args.fonte,
                    arquivo=args.arquivo, sem_ia=args.sem_ia)


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        import traceback
        log("ERRO INESPERADO:\n" + traceback.format_exc())
        sys.exit(1)
